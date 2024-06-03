import nested_admin
import openpyxl
import csv
from django.contrib import admin
from .models import *
from django.utils.html import format_html
from django.http import HttpResponse
from .utils import *

class DiagnosticPlanInline(nested_admin.NestedStackedInline):
    model = DiagnosticPlan
    extra = 0

class DiagnosticLevelInline(nested_admin.NestedStackedInline):
    model = DiagnosticLevel
    inlines = [DiagnosticPlanInline]
    extra = 0

@admin.register(Category)
class CategoryAdmin(nested_admin.NestedModelAdmin):
    inlines = [DiagnosticLevelInline]

class AnswerInline(nested_admin.NestedStackedInline):
    model = Answer
    extra = 0

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        field = super().formfield_for_dbfield(db_field, request, **kwargs)
        if db_field.name == 'value':
            field.widget.attrs['min'] = 0
            field.help_text = 'Value for graphics'
        return field

class QuestionInline(nested_admin.NestedStackedInline):
    model = Question
    inlines = [AnswerInline]
    extra = 0

@admin.register(Form)
class FormAdmin(nested_admin.NestedModelAdmin):
    inlines = [QuestionInline]

@admin.action(description='Exportar seleccionados a CSV')
def export_as_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=completed_forms.csv'
    writer = csv.writer(response)

    if not queryset.exists():
        writer.writerow(['No hay datos para mostrar.'])
        return response

    forms_by_title = {}
    for obj in queryset:
        forms_by_title.setdefault(obj.form_title, []).append(obj)

    for form_title, forms in forms_by_title.items():
        writer.writerow([form_title])
        ordered_answer_questions = get_ordered_answer_questions(forms)
        # Aquí construimos los headers basados en todas las posibles preguntas y categorías
        headers = ['User', 'Form Title', 'Created At'] + list(info_questions.values()) + ordered_answer_questions

        # Añadimos las columnas para cada categoría promedio
        all_category_averages = set()
        for form in forms:
            category_averages = calculate_category_averages(form.content.get('answers', []))
            all_category_averages.update(f"Autodiagnóstico {avg['category']['name']}" for avg in category_averages)
        headers += list(all_category_averages)

        writer.writerow(headers)

        for obj in forms:
            category_averages = calculate_category_averages(obj.content.get('answers', []))
            row = build_row(obj, ordered_answer_questions, category_averages)
            writer.writerow(row)

    return response

@admin.action(description='Exportar seleccionados a Excel')
def export_as_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=completed_forms.xlsx'
    wb = openpyxl.Workbook()

    if not queryset.exists():
        ws = wb.active
        ws.append(['No hay datos para mostrar.'])
        wb.save(response)
        return response

    forms_by_title = {}
    for obj in queryset:
        forms_by_title.setdefault(obj.form_title, []).append(obj)

    for form_title, forms in forms_by_title.items():
        ws = wb.create_sheet(title=form_title[:31])
        ordered_answer_questions = get_ordered_answer_questions(forms)
        # Aquí construimos los headers basados en todas las posibles preguntas y categorías
        headers = ['User', 'Form Title', 'Created At'] + list(info_questions.values()) + ordered_answer_questions

        # Añadimos las columnas para cada categoría promedio
        all_category_averages = set()
        for form in forms:
            category_averages = calculate_category_averages(form.content.get('answers', []))
            all_category_averages.update(f"Autodiagnóstico {avg['category']['name']}" for avg in category_averages)
        headers += list(all_category_averages)

        ws.append(headers)

        for obj in forms:
            category_averages = calculate_category_averages(obj.content.get('answers', []))
            row = build_row(obj, ordered_answer_questions, category_averages)
            ws.append(row)

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(response)
    return response

@admin.register(CompletedForm)
class CompletedFormAdmin(admin.ModelAdmin):
    list_display = ['form_title', 'user', 'email', 'created_at']
    readonly_fields = ('user', 'email', 'form_title', 'created_at', 'content_pretty')
    actions = [export_as_csv, export_as_excel]

    def content_pretty(self, instance):
        answers = instance.content.get('answers', [])
        html = "<table><tr><th>Question</th><th>Category</th><th>Answer</th></tr>"
        for answer in answers:
            question_text = answer.get('questionText', 'N/A')
            category_name = answer.get('category', {}).get('name', 'N/A')
            answer_text = answer.get('answerText', 'N/A')
            html += f"<tr><td>{question_text}</td><td>{category_name}</td><td>{answer_text}</td></tr>"
        html += "</table>"
        return format_html(html)